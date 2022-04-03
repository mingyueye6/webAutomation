import datetime
import json
import os

from openpyxl import load_workbook
import xlrd


def xlrdOperationExcel(file_name="cases.xlsx"):
    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(base_path, 'datas', file_name)
    data = []
    if file_name[-4:] == ".xls":
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        for i in range(1, ws.nrows):
            values = ws.row_values(i)
            dic = {}
            dic["is_run"] = values[1]
            dic["url"] = values[2]
            dic["request_type"] = values[3]
            dic["request_header"] = json.loads(values[4]) if values[4] else None
            dic["request_data"] = json.loads(values[5]) if values[5] else None
            dic["request_cookie"] = values[6]
            dic["response_assert"] = json.loads(values[7]) if values[7] else None
            dic["response_data"] = json.loads(values[8]) if values[8] else None
            data.append([values[0], dic])
    else:
        wb = load_workbook(file_path)
        ws = wb.worksheets[0]
        for row in range(2, ws.max_row + 1):
            values = []
            for i in ws[row]:
                values.append(i.value)
            dic = {}
            dic["is_run"] = values[1]
            dic["url"] = values[2]
            dic["request_type"] = values[3]
            dic["request_header"] = json.loads(values[4]) if values[4] else None
            dic["request_data"] = json.loads(values[5]) if values[5] else None
            dic["request_cookie"] = values[6]
            dic["response_assert"] = json.loads(values[7]) if values[7] else None
            dic["response_data"] = json.loads(values[8]) if values[8] else None
            data.append([values[0], dic])
    return data


if __name__ == '__main__':
    excel = xlrdOperationExcel("cases.xls")
    # excel = xlrdOperationExcel()
    print(excel)
